import io

import pytest

import app


SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"


def test_requested_substance_synonyms():
    assert "Кальций карбонат синтетический" in app.SUBSTANCE_NAMES["3119"]
    assert app.SUBSTANCE_NAMES["0152"] == (
        "Натрий хлорид (Натриевая соль соляной кислоты)"
    )


def test_resolves_position_directly_or_from_legacy_code():
    assert app.resolve_rate_position("(14) Взвешенные вещества") == "14"
    assert app.resolve_rate_position("2902 Взвешенные вещества") == "14"
    assert app.resolve_rate_position("2907 Пыль неорганическая") == "51"
    assert app.resolve_rate_position("2908 Пыль неорганическая") == "51"
    assert app.resolve_rate_position("2909 Пыль неорганическая") == "51"
    assert app.resolve_rate_position("(999) Неизвестное вещество") is None
    assert app.resolve_rate_position("9999 Неизвестное вещество") is None


def test_additional_code_catalog_is_disjoint_and_complete():
    records = app.ADDITIONAL_SUBSTANCES_2909
    assert len(records) == 693
    assert set(records).isdisjoint(app.CODE_TO_POSITION_2909)
    assert {
        group: sum(record["group"] == group for record in records.values())
        for group in range(1, 5)
    } == {1: 4, 2: 653, 3: 28, 4: 8}


def test_resolves_additional_codes_without_matching_by_name():
    assert app.resolve_rate_position("0385 Фтористые соединения") == "67"
    assert app.resolve_rate_position("0102 Алюминия гидроксид") == "14"
    assert app.resolve_rate_position("0124 Кадмий нитрат") == "28"
    assert app.resolve_rate_position("0402 Бутан") == "76"
    assert app.resolve_rate_position("0122 Железо трихлорид") == "25"
    assert app.resolve_rate_position("0242 Железо пентакарбонил") is None
    assert app.resolve_rate_position("Кадмий нитрат") is None


def test_explicit_position_has_priority_and_code_must_start_the_line():
    assert app.resolve_rate_position("(14) 0124 Кадмий нитрат") == "14"
    assert app.extract_code_from_substance("Вещество 0124") is None
    assert app.resolve_rate_position("Вещество 0124") is None


def test_direct_position_and_legacy_code_use_the_same_rate():
    dataframe = app._build_emissions_dataframe(
        [
            ("(14) Взвешенные вещества", {2026: 1.0}),
            ("2902 Взвешенные вещества", {2026: 1.0}),
        ],
        (2026,),
    )

    assert dataframe["Позиция 2909-р"].tolist() == ["14", "14"]
    assert dataframe["Код вещества"].tolist() == [None, "2902"]
    rates = app.PAYMENT_RATES_BY_YEAR["2026"]["rates_by_position"]
    assert [
        app.find_rate_by_position(position, rates)
        for position in dataframe["Позиция 2909-р"]
    ] == [65.5, 65.5]


def test_payment_rates_cover_2026_to_2030():
    expected_2902_rates = {
        "2026": 65.5,
        "2027": 196.6,
        "2028": 327.7,
        "2029": 655.3,
        "2030": 1310.6,
    }

    assert set(app.PAYMENT_RATES_BY_YEAR) == set(expected_2902_rates)
    assert app.AVAILABLE_RATE_YEARS == ["2030", "2029", "2028", "2027", "2026"]

    for year, expected_rate in expected_2902_rates.items():
        rates = app.PAYMENT_RATES_BY_YEAR[year]["rates_by_position"]
        assert len(rates) == 199
        assert set(rates) == {str(position) for position in range(1, 200)}
        assert rates["14"] == expected_rate

    rates_2026 = app.PAYMENT_RATES_BY_YEAR["2026"]["rates_by_position"]
    assert rates_2026["8"] == 9_829_531.5
    assert rates_2026["23"] == 21_144_530_000.0


def test_pyrene_position_and_rates():
    assert len(app.POSITION_TO_RATE_CODES_2909) == 199
    assert app.POSITION_TO_RATE_CODES_2909["100"] == ("0722",)
    assert app.SUBSTANCE_NAMES["0722"] == "Пирен (бензо(d,e,f)фенантрен)"

    expected_rates = {
        "2026": 9_829.5,
        "2027": 29_488.6,
        "2028": 49_147.7,
        "2029": 98_295.3,
        "2030": 196_590.6,
    }
    for year, expected_rate in expected_rates.items():
        rates = app.PAYMENT_RATES_BY_YEAR[year]["rates_by_position"]
        assert rates["100"] == expected_rate


def _xls(rows):
    xml_rows = []
    for row in rows:
        cells = []
        for index, value in sorted(row.items()):
            cells.append(
                f'<Cell ss:Index="{index}"><Data ss:Type="String">{value}</Data></Cell>'
            )
        xml_rows.append(f"<Row>{''.join(cells)}</Row>")
    content = (
        '<?xml version="1.0"?>'
        f'<Workbook xmlns="{SS_NS}" xmlns:ss="{SS_NS}">'
        '<Worksheet ss:Name="Page1"><Table>'
        f"{''.join(xml_rows)}"
        "</Table></Worksheet></Workbook>"
    )
    return io.BytesIO(content.encode("utf-8"))


def test_parses_object_format():
    uploaded = _xls([
        {
            2: (
                "Нормативы выбросов загрязняющих веществ от стационарных ИЗАВ "
                "в атмосферный воздух  по объекту ОНВ."
            )
        },
        _year_headers(5),
        _annual_units(5),
        {
            1: "1",
            2: "0155  Натрия карбонат",
            **_annual_values(6, [0.000036] * 8),
        },
        {
            1: "2",
            2: "0301  Азота диоксид",
            **_annual_values(6, [8.369627] * 8),
        },
    ])

    result = app.parse_emissions_xls(uploaded)

    assert result.success
    assert result.dataframe["Код вещества"].tolist() == ["0155", "0301"]
    assert result.dataframe["Позиция 2909-р"].tolist() == ["34", "1"]
    assert result.dataframe["Валовый выброс, т/год"].tolist() == pytest.approx(
        [0.000036, 8.369627]
    )


def test_parses_source_format_from_substance_totals_only():
    title = (
        "Нормативы выбросов загрязняющих веществ в атмосферный воздух "
        "по конкретным стационарным источникам выбросов и загрязняющим веществам"
    )
    uploaded = _xls([
        {4: title},
        _year_headers(4),
        _annual_units(4),
        {1: "Наименование и код загрязняющего вещества:", 5: "0155 Натрия карбонат"},
        {1: "1", 3: "0006", 5: "0.000020", 6: "ПДВ"},
        {
            2: "Всего по ЗВ",
            **_annual_values(5, [0.000036] * 8),
        },
        # Повторный заголовок страницы не должен создавать запись.
        {4: title},
        _year_headers(4),
        _annual_units(4),
        {1: "Наименование и код загрязняющего вещества:", 5: "0301 Азота диоксид"},
        {1: "2", 3: "0008", 5: "0.060800", 6: "ПДВ"},
        {1: "3", 3: "0009", 5: "5.913600", 6: "ПДВ"},
        {
            2: "Всего по ЗВ",
            **_annual_values(5, [8.369627] * 8),
        },
    ])

    result = app.parse_emissions_xls(uploaded)

    assert result.success
    assert result.dataframe["Код вещества"].tolist() == ["0155", "0301"]
    assert result.dataframe["Валовый выброс, т/год"].tolist() == pytest.approx(
        [0.000036, 8.369627]
    )
    assert len(result.dataframe) == 2


def test_rejects_unknown_spreadsheetml_format():
    result = app.parse_emissions_xls(_xls([{1: "Другая таблица"}]))

    assert not result.success
    assert result.error_category == "unsupported_format"
    assert "Не удалось определить вид таблицы" in result.user_message

class _FakeCell:
    def __init__(self, value):
        self.value = value
        self.ctype = app.xlrd.XL_CELL_EMPTY if value is None else app.xlrd.XL_CELL_TEXT


class _FakeSheet:
    def __init__(self, rows):
        self._rows = rows
        self.nrows = len(rows)
        self.ncols = max(len(row) for row in rows)

    def cell(self, row_idx, col_idx):
        value = self._rows[row_idx][col_idx] if col_idx < len(self._rows[row_idx]) else None
        return _FakeCell(value)


class _FakeWorkbook:
    def __init__(self, rows):
        self.nsheets = 1
        self._sheet = _FakeSheet(rows)

    def sheet_by_index(self, index):
        assert index == 0
        return self._sheet


def test_routes_binary_xls_to_common_object_parser(monkeypatch):
    title = (
        "Нормативы выбросов загрязняющих веществ от стационарных ИЗАВ "
        "в атмосферный воздух по объекту ОНВ."
    )
    year_row = [None] * 28
    units_row = [None] * 28
    for offset, year in enumerate(range(2026, 2034)):
        header_index = 4 + offset * 3
        year_row[header_index] = f"{year} год"
        units_row[header_index] = "г/с"
        units_row[header_index + 1] = "т/г"
        units_row[header_index + 2] = "ПДВ/ВРВ"
    rows = [
        [None, title],
        year_row,
        units_row,
        [1.0, "0155 Натрия карбонат", None, None, None, 0.000036],
        [2.0, "0301 Азота диоксид", None, None, None, 8.369627],
    ]
    monkeypatch.setattr(
        app.xlrd,
        "open_workbook",
        lambda **kwargs: _FakeWorkbook(rows),
    )
    uploaded = io.BytesIO(app.BINARY_XLS_SIGNATURE + b"binary payload")

    result = app.parse_emissions_xls(uploaded)

    assert result.success
    assert result.dataframe["Код вещества"].tolist() == ["0155", "0301"]
    assert result.dataframe["Валовый выброс, т/год"].tolist() == pytest.approx(
        [0.000036, 8.369627]
    )


def test_reports_damaged_binary_xls(monkeypatch):
    def raise_xlrd_error(**kwargs):
        raise app.xlrd.XLRDError("damaged")

    monkeypatch.setattr(app.xlrd, "open_workbook", raise_xlrd_error)
    result = app.parse_emissions_xls(
        io.BytesIO(app.BINARY_XLS_SIGNATURE + b"damaged")
    )

    assert not result.success
    assert result.error_category == "invalid_binary_xls"
    assert "codec" not in result.user_message.lower()


def test_rejects_unknown_file_container():
    result = app.parse_emissions_xls(io.BytesIO(b"not an xls file"))

    assert not result.success
    assert result.error_category == "unsupported_format"


@pytest.mark.parametrize(
    ("value", "expected"),
    [("0,000036", 0.000036), ("8.369627", 8.369627), ("3,60e-05", 0.000036)],
)
def test_parses_supported_xls_number_formats(value, expected):
    assert app._parse_xls_number(value) == pytest.approx(expected)


@pytest.mark.parametrize(
    ("emission", "rate"),
    [
        (None, 10.0),
        (float("nan"), 10.0),
        ("нет данных", 10.0),
        (1.0, None),
        (1.0, "нет данных"),
    ],
)
def test_calculate_payment_returns_none_for_invalid_inputs(emission, rate):
    assert app.calculate_payment(emission, rate, 1, 1) is None



def _year_headers(first_column):
    return {
        first_column + offset * 3: f"{year} год"
        for offset, year in enumerate(range(2026, 2034))
    }


def _annual_units(first_column):
    units = {}
    for offset in range(8):
        first_unit_column = first_column + offset * 3
        units.update({
            first_unit_column: "г/с",
            first_unit_column + 1: "т/г",
            first_unit_column + 2: "ПДВ/\nВРВ",
        })
    return units


def _annual_values(first_tons_column, values):
    return {
        first_tons_column + offset * 3: value
        for offset, value in enumerate(values)
        if value is not None
    }


def test_parses_all_eight_yearly_object_norms():
    title = (
        "Нормативы выбросов загрязняющих веществ от стационарных ИЗАВ "
        "в атмосферный воздух по объекту ОНВ."
    )
    annual_values = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8]
    uploaded = _xls([
        {2: title},
        _year_headers(5),
        _annual_units(5),
        {
            1: "1",
            2: "0155 Натрия карбонат",
            **_annual_values(6, annual_values),
        },
    ])

    result = app.parse_emissions_xls(uploaded)

    assert result.success
    assert result.years == tuple(range(2026, 2034))
    assert result.warnings == ()
    assert result.dataframe.loc[0, "Валовый выброс, т/год"] == pytest.approx(0.1)
    assert list(
        result.dataframe.loc[0, "Нормативы по годам"].values()
    ) == pytest.approx(annual_values)


def test_parses_all_eight_yearly_source_totals():
    title = (
        "Нормативы выбросов загрязняющих веществ в атмосферный воздух "
        "по конкретным стационарным источникам выбросов и загрязняющим веществам"
    )
    annual_values = [1.0, 1.2, 1.4, 1.6, 1.8, 2.0, 2.2, 2.4]
    uploaded = _xls([
        {4: title},
        _year_headers(4),
        _annual_units(4),
        {1: "Наименование и код загрязняющего вещества:", 5: "0301 Азота диоксид"},
        {2: "Всего по ЗВ", **_annual_values(5, annual_values)},
    ])

    result = app.parse_emissions_xls(uploaded)

    assert result.success
    assert result.warnings == ()
    assert result.dataframe.loc[0, "Нормативы по годам"] == pytest.approx(
        dict(zip(range(2026, 2034), annual_values))
    )


def test_keeps_missing_years_empty_and_reports_warnings():
    title = (
        "Нормативы выбросов загрязняющих веществ от стационарных ИЗАВ "
        "в атмосферный воздух по объекту ОНВ."
    )
    row = {
        1: "1",
        2: "0155 Натрия карбонат",
        **_annual_values(6, [0.1, None, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8]),
    }
    row[12] = "нет данных"
    uploaded = _xls([
        {2: title},
        _year_headers(5),
        _annual_units(5),
        row,
    ])

    result = app.parse_emissions_xls(uploaded)

    norms = result.dataframe.loc[0, "Нормативы по годам"]
    assert result.success
    assert norms[2026] == pytest.approx(0.1)
    assert norms[2027] is None
    assert norms[2028] is None
    assert len(result.warnings) == 2



def test_rejects_moved_annual_column():
    title = (
        "Нормативы выбросов загрязняющих веществ от стационарных ИЗАВ "
        "в атмосферный воздух по объекту ОНВ."
    )
    headers = _year_headers(5)
    headers[9] = headers.pop(8)
    uploaded = _xls([
        {2: title},
        headers,
        _annual_units(5),
        {
            1: "1",
            2: "0155 Натрия карбонат",
            **_annual_values(6, [0.1] * 8),
        },
    ])

    result = app.parse_emissions_xls(uploaded)

    assert not result.success
    assert result.error_category == "modified_structure"
    assert result.user_message == app.MODIFIED_STRUCTURE_MESSAGE


def test_rejects_deleted_year_header():
    title = (
        "Нормативы выбросов загрязняющих веществ от стационарных ИЗАВ "
        "в атмосферный воздух по объекту ОНВ."
    )
    headers = _year_headers(5)
    headers.pop(11)
    uploaded = _xls([{2: title}, headers, _annual_units(5)])

    result = app.parse_emissions_xls(uploaded)

    assert not result.success
    assert result.error_category == "modified_structure"


def test_rejects_missing_tons_unit_column():
    title = (
        "Нормативы выбросов загрязняющих веществ от стационарных ИЗАВ "
        "в атмосферный воздух по объекту ОНВ."
    )
    units = _annual_units(5)
    units.pop(9)
    uploaded = _xls([{2: title}, _year_headers(5), units])

    result = app.parse_emissions_xls(uploaded)

    assert not result.success
    assert result.error_category == "modified_structure"


def test_rejects_modified_structure_on_repeated_page():
    title = (
        "Нормативы выбросов загрязняющих веществ от стационарных ИЗАВ "
        "в атмосферный воздух по объекту ОНВ."
    )
    repeated_headers = _year_headers(5)
    repeated_headers[12] = repeated_headers.pop(11)
    uploaded = _xls([
        {2: title},
        _year_headers(5),
        _annual_units(5),
        {
            1: "1",
            2: "0155 Натрия карбонат",
            **_annual_values(6, [0.1] * 8),
        },
        repeated_headers,
        _annual_units(5),
    ])

    result = app.parse_emissions_xls(uploaded)

    assert not result.success
    assert result.error_category == "modified_structure"


def _test_rates_by_year():
    return {
        str(year): {"rates_by_position": {"14": float(year - 2025)}}
        for year in range(2026, 2031)
    }


def test_selected_year_uses_matching_norm_and_rate():
    source = app.pd.DataFrame({
        "Наименование вещества": ["Взвешенные вещества"],
        "Позиция 2909-р": ["14"],
        "Валовый выброс, т/год": [1.0],
        "Нормативы по годам": [{2026: 1.0, 2027: 2.0, 2028: 3.0}],
    })

    calculated = app.calculate_selected_year_dataframe(
        source,
        "2028",
        _test_rates_by_year()["2028"]["rates_by_position"],
        1,
        25,
    )

    assert calculated.loc[0, "Валовый выброс, т/год"] == pytest.approx(3.0)
    assert calculated.loc[0, "Ставка платы, руб."] == pytest.approx(3.0)
    assert calculated.loc[0, "Сумма платы, руб/год"] == pytest.approx(225.0)


def test_builds_single_year_export_with_total():
    df = app.pd.DataFrame({
        "Наименование вещества": ["Взвешенные вещества"],
        "Валовый выброс, т/год": [2.0],
        "Ставка платы, руб.": [5.0],
        "Сумма платы, руб/год": [10.0],
    })

    single = app.build_single_year_export_dataframe(df)

    assert single.columns.tolist() == [
        "Наименование вещества",
        "Валовый выброс, т/год",
        "Ставка платы, руб.",
        "Сумма платы, руб/год",
    ]
    assert single.iloc[-1]["Наименование вещества"] == "ИТОГО:"
    assert single.iloc[-1]["Валовый выброс, т/год"] == pytest.approx(2.0)
    assert single.iloc[-1]["Сумма платы, руб/год"] == pytest.approx(10.0)


def test_builds_yearly_export_with_year_specific_and_fallback_rates():
    years = tuple(range(2026, 2034))
    norms = {year: float(offset + 1) for offset, year in enumerate(years)}
    df = app.pd.DataFrame({
        "Наименование вещества": ["0155 Натрия карбонат"],
        "Позиция 2909-р": ["14"],
        "Нормативы по годам": [norms],
    })

    yearly = app.build_yearly_export_dataframe(
        df, years, _test_rates_by_year(), 1, 25
    )

    expected_columns = ["Наименование вещества"]
    for year in years:
        expected_columns.extend([
            f"Норматив {year}, т/год",
            f"Ставка {year}, руб./т",
            f"Сумма платы {year}, руб/год",
        ])
    assert yearly.columns.tolist() == expected_columns
    assert yearly.loc[0, "Норматив 2027, т/год"] == pytest.approx(2.0)
    assert yearly.loc[0, "Ставка 2027, руб./т"] == pytest.approx(2.0)
    assert yearly.loc[0, "Сумма платы 2027, руб/год"] == pytest.approx(100.0)
    assert yearly.loc[0, "Ставка 2033, руб./т"] == pytest.approx(5.0)
    assert yearly.loc[1, "Сумма платы 2033, руб/год"] == pytest.approx(1000.0)


def test_sums_available_values_when_norm_is_missing():
    years = tuple(range(2026, 2034))
    norms = {year: 1.0 for year in years}
    norms[2029] = None
    df = app.pd.DataFrame({
        "Наименование вещества": ["0155 Натрия карбонат"],
        "Позиция 2909-р": ["14"],
        "Нормативы по годам": [norms],
    })

    yearly = app.build_yearly_export_dataframe(
        df, years, _test_rates_by_year(), 1, 1
    )

    assert app.pd.isna(yearly.loc[0, "Норматив 2029, т/год"])
    assert yearly.loc[0, "Ставка 2029, руб./т"] == pytest.approx(4.0)
    assert app.pd.isna(yearly.loc[0, "Сумма платы 2029, руб/год"])
    assert yearly.loc[1, "Норматив 2029, т/год"] == 0.0
    assert yearly.loc[1, "Сумма платы 2029, руб/год"] == 0.0


def test_missing_rate_keeps_payment_blank_and_excludes_it_from_total():
    years = tuple(range(2026, 2034))
    norms = {year: 1.0 for year in years}
    df = app.pd.DataFrame({
        "Наименование вещества": ["Вещество 1", "Вещество 2"],
        "Позиция 2909-р": ["14", None],
        "Нормативы по годам": [norms, norms],
    })

    yearly = app.build_yearly_export_dataframe(
        df, years, _test_rates_by_year(), 1, 1
    )

    assert app.pd.isna(yearly.loc[1, "Ставка 2026, руб./т"])
    assert app.pd.isna(yearly.loc[1, "Сумма платы 2026, руб/год"])
    assert yearly.loc[2, "Сумма платы 2026, руб/год"] == 1.0


def test_excel_output_contains_two_styled_sheets_and_fallback_note():
    from openpyxl import load_workbook

    years = tuple(range(2026, 2034))
    norms = {year: 1.0 for year in years}
    source_df = app.pd.DataFrame({
        "Наименование вещества": ["0155 Натрия карбонат"],
        "Позиция 2909-р": ["14"],
        "Валовый выброс, т/год": [1.0],
        "Ставка платы, руб.": [1.0],
        "Сумма платы, руб/год": [1.0],
        "Нормативы по годам": [norms],
    })
    single_sheet = app.build_single_year_export_dataframe(source_df)
    yearly_sheet = app.build_yearly_export_dataframe(
        source_df,
        years,
        _test_rates_by_year(),
        1,
        1,
    )

    workbook = load_workbook(
        app.create_excel_output(
            single_sheet,
            yearly_sheet,
            years,
            _test_rates_by_year(),
            "2028",
        ),
        data_only=False,
    )

    assert workbook.sheetnames == ["Расчет платы", "Расчёт по годам"]
    single = workbook["Расчет платы"]
    assert single["A1"].value == "Наименование вещества"
    assert single["B1"].value == "Валовый выброс, т/год"
    assert single["C1"].value == "Ставка платы, руб."
    assert single["D1"].value == "Сумма платы, руб/год"
    assert single.freeze_panes == "B2"
    single_note_row = len(single_sheet) + 3
    single_note = single.cell(row=single_note_row, column=1).value
    assert single_note == (
        "Примечание: расчёт выполнен по нормативам и ставкам за 2028 год."
    )
    assert f"A{single_note_row}:D{single_note_row}" in {
        str(cell_range) for cell_range in single.merged_cells.ranges
    }

    yearly = workbook["Расчёт по годам"]
    assert yearly["A1"].value == "Наименование вещества"
    assert yearly["B1"].value == "Норматив 2026, т/год"
    assert yearly["C1"].value == "Ставка 2026, руб./т"
    assert yearly["D1"].value == "Сумма платы 2026, руб/год"
    assert yearly.freeze_panes == "B2"

    payment_columns = set(range(4, 26, 3))
    rate_columns = set(range(3, 26, 3))
    total_row = len(yearly_sheet) + 1
    for row in yearly.iter_rows(min_row=1, max_row=total_row):
        for cell in row:
            if cell.column in payment_columns:
                assert cell.fill.fill_type == "solid"
                assert cell.fill.fgColor.rgb.endswith("E4F1E8")
            else:
                assert cell.fill.fill_type is None

            assert cell.border.left.style == "thin"
            assert cell.border.right.style == "thin"
            assert cell.border.bottom.style == "thin"
            expected_top = "medium" if cell.row == total_row else "thin"
            assert cell.border.top.style == expected_top

    note_row = len(yearly_sheet) + 3
    note = yearly.cell(row=note_row, column=1).value
    assert note == (
        "Примечание: для расчёта платы за 2031–2033 годы условно применены "
        "ставки платы, установленные на 2030 год, в связи с отсутствием "
        "утверждённых ставок на указанный период."
    )
    assert f"A{note_row}:Y{note_row}" in {
        str(cell_range) for cell_range in yearly.merged_cells.ranges
    }
