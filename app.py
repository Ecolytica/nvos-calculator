import streamlit as st
import pandas as pd
import io
import json
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import numpy as np
import xlrd

from analytics import (
    flush_events,
    queue_event,
    render_feedback,
    render_page_view,
    render_sidebar_about,
    render_social_button,
)

# Настройка страницы
st.set_page_config(
    page_title="Расчет платы за выбросы в атмосферу",
    layout="wide",
    # Streamlit keeps the sidebar open on wide screens and collapses it on
    # narrow ones, where it becomes an accessible slide-out panel.
    initial_sidebar_state="auto",
)

def load_ecolytica_styles():
    """Подключает локальную тему Ecolytica."""
    css_path = Path(__file__).parent / "assets" / "ecolytica.css"
    st.markdown(f"<style>{css_path.read_text(encoding='utf-8')}</style>", unsafe_allow_html=True)


load_ecolytica_styles()
render_page_view()

st.markdown(
    """
    <div class="eco-page-head">
      <div class="eco-eyebrow">ПЛАТА ЗА НВОС</div>
      <h1>Калькулятор платы за выбросы в атмосферу</h1>
    </div>
    <div class="eco-section-label">Исходные данные</div>
    """,
    unsafe_allow_html=True,
)

# Названия веществ не зависят от года; числовые ставки загружаются из JSON.
SUBSTANCE_NAMES = {
    '0301': 'Азота диоксид (двуокись азота; пероксид азота)',
    '0304': 'Азота оксид (азот (II) оксид; азот монооксид)',
    '0302': 'Азотная кислота (по молекуле HNO3)',
    '0303': 'Аммиак (азота гидрид)',
    '0305': 'Аммиачная селитра (аммоний нитрат; аммоний азотнокислый)',
    '0231': 'Барий и его соли (ацетат, нитрат, нитрит, хлорид) /в пересчете на барий/',
    '0104': 'Барий карбонат (барий углекислый) /в пересчете на барий/',
    '0703': 'Бенз(а)пирен',
    '0109': 'Бериллий и его соединения /в пересчете на бериллий/',
    '0308': 'Борная кислота (ортоборная кислота; орто-борная кислота; бор тригидрооксид)',
    '0110': 'Ванадия пяти оксид (диванадий пентоксид (пыль); ванадиевый ангидрид)',
    '0008': 'Взвешенные частицы PM10',
    '0010': 'Взвешенные частицы PM2,5',
    '2902': 'Взвешенные вещества (разнородные по составу твердые частицы, содержащиеся в выбросах загрязняющих веществ и не поименованные в настоящем разделе)',
    '0313': 'Водород бромистый (гидробромид)',
    '0314': 'Водород мышьяковистый (арсин)',
    '0315': 'Водород фосфористый (фосфин; гидроген фосфид)',
    '0317': 'Водород цианистый (гидроцианид; синильная кислота; нитрил муравьиной кислоты; цианистоводородная кислота; формонитрил)',
    '0369': 'Гексафторид серы (сера гексафторид (ОС-6-11); (ОСС-6-11) сера фторид)',
    '0101': 'диАлюминий триоксид /в пересчете на алюминий/',
    '0123': 'диЖелезо триоксид (железа оксид; железо сесквиоксид) /в пересчете на железо/',
    '0158': 'диНатрий сульфат (натрий сернокислый; динатриевая соль серной кислоты; динатрий сернокислый)',
    '3620': 'Диоксины (полихлорированные дибензо-n-диоксины и дибензофураны) /в пересчете на 2,3,7,8-тетрахлордибензо-1,4-диоксин/',
    '0119': 'Диэтилртуть /в пересчете на ртуть/',
    '0122': 'Железа трихлорид (железо (III) хлорид; железо перхлорид; железо хлорное) /в пересчете на железо/',
    '2903': 'Зола сланцевая',
    '2904': 'Зола ТЭС мазутная /в пересчете на ванадий/',
    '0133': 'Кадмий и его соединения (кадмий дийодид (йодистый кадмий); кадмий динитрат (кадмий азотнокислый тетрагидрат); кадмий дихлорид (хлористый кадмий); кадмий оксид; кадмий сульфат (кадмий сульфат октагидрат) /в пересчете на кадмий/',
    '0126': 'Калий хлорид (калиевая соль соляной кислоты)',
    '3119': 'Кальций карбонат (Кальций карбонат синтетический; кальций углекислый; кальциевая соль карбоновой кислоты (1:1)',
    '0128': 'Кальций оксид (кальций окись)',
    '0150': 'Натрий гидроксид (натр едкий)',
    '0152': 'Натрий хлорид (Натриевая соль соляной кислоты)',
    '0155': 'Карбонат натрия (динатрий карбонат; натрий углекислый; натриевая соль угольной кислоты)',
    '1551': 'Кислота терефталевая (бензол-1,4-дикарбоновая кислота; п-фталевая кислота; бензол-п-дикарбоновая кислота)',
    '0134': 'Кобальт и его соединения (кобальт; кобальт оксид (кобальт окись, кобальт монооксид, кобальт (II) оксид); кобальт сульфат (кобальт моносульфат гептагидрат); диацетат кобальта (II) (кобальт (II) уксуснокислый тетрагидрат)) /в пересчете на кобальт/',
    '0164': 'Никель, оксид никеля /в пересчете на никель/',
    '0165': 'Никель растворимые соли /в пересчете на никель/',
    '0138': 'Магний оксид (окись магния)',
    '0143': 'Марганец и его соединения /в пересчете на марганец (IV) оксид/',
    '0146': 'Медь и ее соединения (медь оксид (медь окись; тенорит); медь сульфат (медь сернокислая; медная соль серной кислоты); медь сульфит (1:1); медь хлорид (монохлорид меди; хлористая медь); медь дихлорид (медь (II) хлорид) /в пересчете на медь/',
    '0410': 'Метан',
    '1715': 'Метилмеркаптан (метантиол)',
    '1728': 'Этилмеркаптан (этантиол; меркаптоэтан; этилсульфагидрат; этилгидросульфат; тиоэтиловый спирт; тиоэтанол)',
    '0325': 'Мышьяк и его соединения /в пересчете на мышьяк/, кроме водорода мышьяковистого',
    '0326': 'Озон (трехатомный кислород)',
    '2930': 'Пыль абразивная (корунд белый, монокорунд)',
    '2934': 'Пыль аминопластов',
    '2931': 'Пыль асбестосодержащая (с содержанием хризотиласбеста до 10 процентов) (по асбесту)',
    '3749': 'Пыль каменного угля',
    '2909': 'Пыль неорганическая с содержанием кремния 20 - 70, а также более 70 процентов',
    '2908': 'Пыль неорганическая с содержанием кремния менее 20',
    '2907': 'Пыль неорганическая с содержанием кремния более 70 процентов',
    '2953': 'Пыль фенопластов резольного типа (Э2-330-02, У2-301-07)',
    '0183': 'Ртуть и ее соединения, кроме диэтилртути (в том числе: ртуть оксид; ртуть хлорид; ртуть дихлорид; диацетат ртути; ртуть амидохлорид; ртуть дийодид; ртуть динитрат гидрат; ртуть нитрат дигидрат) /в пересчете на ртуть/',
    '0184': 'Свинец и его соединения, кроме тетраэтилсвинца /в пересчете на свинец/',
    '0333': 'Сероводород (дигидросульфид; водород сернистый; дигидросульфид; гидросульфид)',
    '0334': 'Сероуглерод (углерод сульфид; углерод двусернистый; дитиокарбоновый ангидрид; сульфокарбоновый ангидрид)',
    '0322': 'Серная кислота (по молекуле H2SO4)',
    '0330': 'Серы диоксид',
    '3748': 'Смолистые вещества (возгоны пека) в составе электролизной пыли выбросов производства алюминия',
    '0193': 'Теллура диоксид /в пересчете на теллур/',
    '0192': 'Тетраэтилсвинец',
    '0118': 'Титана диоксид (титан пероксид; титан (IV) оксид)',
    '0328': 'Углерод (пигмент черный или углеродсодержащий аэрозоль (сажа)',
    '0337': 'Углерода оксид (углерод окись; углерод моноокись; угарный газ)',
    '0347': 'Фосген (карбонилдихлорид)',
    '0338': 'Фосфорный ангидрид (дифосфор пентаоксид; фосфор (V) оксид)',
    '0342': 'Фториды газообразные /в пересчете на фтор/: гидрофторид (водород фторид, фторводород); кремний тетрафторид',
    '0344': 'Фториды твердые (фториды неорганические плохо растворимые): алюминия фторид; кальция фторид; натрия гексафторалюминат',
    '0343': 'Фтористый водород, растворимые фториды (фториды неорганические хорошо растворимые): натрия фторид (натрий фтористый); натрия гексафторидсиликат',
    '0349': 'Хлор',
    '0378': 'Хлор диоксид',
    '0316': 'Хлористый водород (гидрохлорид, водород хлорид) /по молекуле HCl/',
    '0203': 'Хром /в пересчете на хрома (VI) оксид/',
    '0207': 'Цинк оксид /в пересчете на цинк/',
    '0205': 'Цинк сульфат /в пересчете на цинк/',
    '0415': 'Углеводороды предельные C1 - C5 (смесь предельных углеводородов C1H4 - C5H12) (исключая метан)',
    '0416': 'Углеводороды предельные C6 - C10 (смесь предельных углеводородов C6H14 - C10H14)',
    '2754': 'Углеводороды предельные C12 - C19 (растворители РПК-240, РПК-280)',
    '0408': 'Циклогексан (гексаметилен; гексагидробензол; бензолгексагидрид)',
    '0501': 'Амилены (смесь изомеров; пентилены)',
    '0502': 'Бутилен (бут-1-ен; альфа-бутилен; 1-бутен; 1-бутилен; этилэтилен; н-бутен)',
    '0503': '1,3-Бутадиен (дивинил; бута-1,3-диен; альфа, гамма-бутадиен; 1-метилаллен; биэтилен; дивинил; винилэтилен; бивинил)',
    '0507': 'Гекс-1-ен (бутилэтилен; альфа-гексилен; 1-н-гексен)',
    '0508': 'Гептен (гепт-1-ен)',
    '0516': '2-Метилбута-1,3-диен (изопентадиен; бета-метилдивинил; гермитерпен; 2-метил-1,3-бутадиен; 2-метилбута-диен-1,3; изопрен)',
    '0521': 'Пропилен (пропен; метилэтилен; пропилен-1; пропен-1)',
    '0526': 'Этилен (этен)',
    '0618': 'Альфа-метилстирол (1-метиэтенил) бензол; 1-метил-1-фенилэтен; 2-фенилпропен-1; изопропенилбензол)',
    '0602': 'Бензол (циклогексатриен; фенилгидрид)',
    '0616': 'Диметилбензол (ксилол) (смесь о-, м-, п- изомеров (метилтолуол))',
    '0612': 'Изопропилбензол (кумол; (1-метилэтил)бензол; 2-фенилпропан)',
    '0621': 'Метилбензол (фенилметан; толуол)',
    '0617': 'Растворитель мебельный (AMP-3) /по толуолу/',
    '0623': '1,3,5-Триметилбензол (мезитилен; триметилбензол симметричный; 3,5-диметилтолуол)',
    '1071': 'Фенол (гидроксибензол; оксибензол; фенилгидроксид; фениловый спирт; моногидроксибензол)',
    '0627': 'Этилбензол (фенилэтан)',
    '0620': 'Этенилбензол (стирол; винилбензол; фенилэтилен)',
    '0711': 'Антрацен',
    '0708': 'Нафталин (нафтален; нафтен)',
    '0722': 'Пирен (бензо(d,e,f)фенантрен)',
    '0716': 'Фенантрен',
    '0810': 'Бромбензол',
    '0813': '1-Бромгептан (гептил бромистый; гептилбромид)',
    '0814': '1-Бромдекан (децил бромистый)',
    '0815': '1-Бром-3-метилбутан (изоамил бромистый; изоамиловый бромид)',
    '0816': '1-Бром-2-метилпропан (изобутил бромистый)',
    '0819': '1-Бромпентан (амил бромистый; амил бромид)',
    '0817': '1-Бромпропан (пропил бромистый)',
    '0818': '2-Бромпропан (изопропил бромистый)',
    '4005': '1,1-Дихлорэтан',
    '0856': '1,2-Дихлорэтан',
    '0858': 'Дихлорфторметан (фреон 21)',
    '0859': 'Дифторхлорметан (фреон 22)',
    '0861': '1,2-Дихлорпропан',
    '0869': 'Метилен хлористый (дихлорметан; метиленхлорид; метан дихлорид; метилен бихлорид; метилен хлорид; метилен дихлорид)',
    '0906': 'Тетрахлорметан (углерод четыреххлористый; углерод тетрахлорид; перхлорметан; тетрахлоруглерод)',
    '0882': 'Тетрахлорэтилен (перхлорэтилен; тетрахлорид этилена; 1,1,2,2-тетрахлорэтилен; тетрахлорэтен)',
    '0883': 'Тетрафторэтилен',
    '0898': 'Трихлорметан (хлороформ)',
    '0902': 'Трихлорэтилен (1-хлор-2,2-дихлорэтилен; этилентрихлорид; ацетилентрихлорид; 1,1,2-трихлорэтилен)',
    '0890': 'Трибромметан (бромоформ; метилтрибромид)',
    '0915': 'Хлорбензол (фенилхлорид)',
    '0932': 'Хлорэтан (этил хлористый; хлорэтил)',
    '0827': 'Хлорэтен (хлорэтилен; винихлорид; винил хлористый; хлорвинил; этиленхлорид)',
    '0930': 'Хлоропрен (2-хлорбута-1,3-диен; полихлорпрен, поли-2-хлор-1,3-бутадиен)',
    '0931': 'Эпихлоргидрин ((хлорметил) оксиран+; 1-хлор-2,3-эпоксипропан; 1-хлорпропеноксид; 3-хлорпропеноксид; глицидилхлорид; хлорметилоксиран)',
    '1043': 'Гексан-1-ол (гексиловый спирт, 1-гексанол; 1-гидроксигексан; амилкарбинол; пентилкарбинол)',
    '1069': 'Гидроксиметилбензол (крезол; смесь изомеров о-, м-, п-) (метилфенол (смесь изомеров); гидрокситолуол (смесь изомеров)',
    '1039': 'Спирт амиловый (пентан-1-ол; н-амиловый спирт; н-пентанол; пентанол-1; бутилкарбинол)',
    '1042': 'Спирт бутиловый (бутан-1-ол)',
    '1048': 'Спирт изобутиловый (2-метилпропан-1-ол; изобутанол; 1-гидроксиметилпропан; 2-метил-1-пропанол; 2-метилпропиловый спирт; изопропилкарбинол)',
    '1050': 'Спирт изооктиловый (2-этилгексанол; 2-этилгексиловый спирт)',
    '1051': 'Спирт изопропиловый (пропан-2-ол; изопропанол; диметилкарбинол; вторичный пропиловый спирт)',
    '1052': 'Спирт метиловый (метанол; карбинол; метиловый спирт; метилгидроксид; моногидроксиметан)',
    '1054': 'Спирт пропиловый (пропан-1-ол; этилкарбинол; 1-оксипропан; пропанол-1; 1-пропанол; н-пропанол; н-пропан-1-ол; 1-гидроксипропан; н-пропиловый спирт)',
    '1061': 'Спирт этиловый (этанол; этиловый спирт; метилкарбинол)',
    '1077': 'Циклогексанол (гексагидрофенол; гексалин; гидроксициклогексан; оксициклогексан; цилогексиловый спирт)',
    '1211': 'Диметиловый эфир терефталевой кислоты (диметилбензол-1,4-дикарбонат; диметил-1,4-бензолдикарбоксилат; диметиловый эфир 1,4-бензолдикарбоновой кислоты; диметиловый эфир терефталевой кислоты)',
    '1103': 'Динил (смесь 25 процентов дифенила и 75 процентов дифенилоксида)',
    '1217': 'Диоктилбензол-1,2-дикарбонат',
    '1105': "Диэтиловый эфир (этоксиэтан; 1,1'-оксибисэтан, оксибис-1,1'-этан, 3-оксапентан, диэтилоксид)",
    '1319': 'Метилаль (диметоксиметан; формаль диметилацеталь)',
    '1140': 'Моноизобутиловый эфир этиленгликоля (бутилцеллозольв; 2-бутоксиэтанол; бутилгликоль; этиленгликоль монобутиловый эфир; монобутиловый эфир этиленгликоля)',
    '1246': 'Этилформиат (муравьиноэтиловый эфир; этилметаноат)',
    '1119': '2-Этоксиэтанол (моноэтиловый эфир этиленгликоля; этилцеллозольв)',
    '1206': 'Бутилакрилат (бутиловый эфир акриловой кислоты; бутилпроп-2-еноат; бутилпропеноат; бутиловый эфир пропеновой кислоты)',
    '1210': 'Бутилацетат (бутиловый эфир уксусной кислоты)',
    '1213': 'Винилацетат (этенилацетат; виниловый эфир уксусной кислоты; этениловый эфир уксусной кислоты; этениловый эфир этановой кислоты; этенилацетат, 1-ацетоксиэтенил)',
    '1225': 'Метилакрилат (метилпроп-2-еноат; метиловый эфир акриловой кислоты; метиловый эфир 2-пропеновой кислоты)',
    '1224': 'Метилацетат (метиловый эфир уксусной кислоты, метилэтаноат, уксуснометиловый эфир)',
    '1240': 'Этилацетат (этиловый эфир уксусной кислоты)',
    '1301': 'Акролеин (проп-2-ен-1-аль; акрилальдегид; акриловый альдегид; альдегид акриловой кислоты)',
    '1310': 'Альдегид масляный (бутаналь; бутальдегид; н-бутиральдегид; бутиловый альдегид)',
    '1317': 'Ацетальдегид (уксусный альдегид)',
    '1314': 'Пропаналь (пропиональдегид, метилацетальдегид)',
    '1325': 'Формальдегид (муравьиный альдегид, оксометан, метиленоксид)',
    '2425': 'Фуран-2-альдегид+ (2-фуральдегид; фурфураль; 2-фурфуральдегид)',
    '1401': 'Ацетон (пропан-2-он; диметилкетон; диметилформальдегид)',
    '1402': 'Ацетофенон (метилфенилкетон; 1-фенилэтанон; фенилметилкетон; ацетилбензол)',
    '1409': 'Метилэтилкетон (бутан-2-он; этилметилкетон; метилацетон)',
    '1405': 'Растворитель древесноспиртовой марки А (ацетоно-эфирный) /контроль по ацетону/',
    '1406': 'Растворитель древесноспиртовой марки Э (эфирно-ацетоновый) /контроль по ацетону/',
    '1411': 'Циклогексанон (циклогексил кетон; кетогексаметилен; пиметинкетон; гексанон)',
    '1505': 'Ангидрид малеиновый (пары, аэрозоль) (дигидрофуран-2,5-дион; малеиновый кислоты ангидрид; цис-1,2-этилендикарбоновой кислоты ангидрид; цис-бутендиовой кислоты ангидрид; 2,5-фурандион; дигидро-2,5-диоксофуран)',
    '1507': 'Ангидрид уксусный (ацетангидрид; этановый ангидрид)',
    '1508': 'Ангидрид фталевый (изобензофуран-1,3-дион; фталевой кислоты ангидрид; 1,3-изобензофурандион; бензол-1,2-дикарбоновой кислоты ангидрид; 1,3-дигидро-1,3-диоксоизобензофуран)',
    '1523': 'Диметилформамид (N,N-диметилформамид; диметиламид муравьиной кислоты; N-формилдиметиламин)',
    '1530': 'Эпсилон-капролактам (гексагидро-2H-азепин-2-он; 4-аминокапроновой кислоты лактам, 2-аминогексиновой кислоты лактам, 2-оксогексаметиленимин, 1,6-гексолактам, 1-аза-2-циклогептанон, 2 кетогексаметиленимин, 6-гексанлактам, 2-пергидроазепинон)',
    '1512': 'Кислота акриловая (проп-2-еновая кислота; этиленкарбоновая кислота)',
    '1519': 'Кислота валериановая (пентановая кислота; 1-бутанкарбоновая кислота; пропилуксусная кислота)',
    '1531': 'Кислота капроновая (гексановая кислота)',
    '1534': 'Кислота масляная (бутановая кислота; этилуксусная кислота, н-бутановая кислота; 1-пропанкарбоновая кислота; пропилмуравьиная кислота)',
    '1546': 'Кислота пропионовая (метилуксусная кислота; этанкарбоновая кислота; этилмуравьиная кислота; карбоксиэтан)',
    '1555': 'Кислота уксусная (этановая кислота; метанкарбоновая кислота)',
    '1537': 'Кислота муравьиная',
    '1601': 'Гидроперекись изопропилбензола (гидроперекись кумола; 1-метил-1-фенилэтилгидропероксид; гидропероксид кумола, кумилгидропероксид; альфа, альфа-диметилбензил-гидропероксид)',
    '1608': 'Пропилена окись (1,2-эпоксипропан; 1,2-пропиленоксид; метилоксиран; альфа-пропиленоксид; метилэтилоксид)',
    '1611': 'Этилена окись (эпоксиэтан; оксиран; этиленоксид)',
    '1702': '1-Бутантиол (н-бутантиол; бутилмеркаптан)',
    '1707': 'Диметилсульфид (метилсульфид; тиобис(метан); метантиометан)',
    '1706': 'Диметилдисульфид (2,3-дитиобутан; (метилдисульфанил)метан)',
    '1720': 'Пропан-1-тиол (пропантиол, пропилмеркаптан)',
    '1716': 'Одорант СПМ-ТУ 51-81-88 (смесь природных меркаптанов с массовым содержанием этантиола 26 - 41 процентов, изопропан-тиола 38 - 47 процентов, втор-бутантиола 7 - 13 процентов) /в пересчете на этилмеркаптан/',
    '1805': 'Анилин (аминобензол; фениламин; бензоламин)',
    '1819': 'Диметиламин',
    '1849': 'Метиламин (аминометан; метанамин; монометиламин)',
    '2418': 'Пиридин (азабензол; азин)',
    '1863': 'Триэтиламин ((диэтиламин)этан)',
    '1905': 'Нитробензол (мононитробензол)',
    '2001': 'Акрилонитрил (проп-2-еннитрил; винил цианистый; нитрил акриловой кислоты; цианоэтилен; винилцианид)',
    '2009': 'N,N-Диметилацетамид (диметиламид уксусной кислоты; ацетилдиметиламин; N,N-диметилэтанамид)',
    '2031': 'Толуилендиизоцианат (диизоцианатметилбензол; метилфенилдиизоцианат; смесь метил-мета-фениловых эфиров изоциановой кислоты; толуолдиизоцианат)',
    '2704': 'Бензин (нефтяной, малосернистый) /в пересчете на углерод/',
    '2705': 'Бензин сланцевый /в пересчете на углерод/',
    '2732': 'Керосин (керосин прямой перегонки; керосин дезодорированный)',
    '2735': 'Минеральное масло (масло минеральное нефтяное): веретенное, машинное, цилиндровое и иные',
    '2748': 'Скипидар /в пересчете на углерод/',
    '2750': 'Сольвент нафта',
    '2752': 'Уайт-спирит',
}

POSITION_CROSSWALK_PATH = (
    Path(__file__).parent / "data" / "position_2909_to_rate_codes.json"
)
ADDITIONAL_CODES_2909_PATH = (
    Path(__file__).parent / "data" / "additional_substance_codes_2909.json"
)
PAYMENT_RATES_PATH = Path(__file__).parent / "data" / "payment_rates.json"


def load_position_crosswalk(path, expected_codes):
    """Загружает соответствия позиций № 2909-р кодам ставок."""
    try:
        catalog = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ValueError(f"файл {path.name} не найден") from exc
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"ошибка JSON в строке {exc.lineno}, столбце {exc.colno}"
        ) from exc

    regulation = catalog.get("regulation") if isinstance(catalog, dict) else None
    positions = catalog.get("positions") if isinstance(catalog, dict) else None
    if not isinstance(regulation, str) or not regulation.strip():
        raise ValueError("не указан источник regulation")
    if not isinstance(positions, dict) or not positions:
        raise ValueError("раздел positions отсутствует или пуст")

    normalized = {}
    code_positions = {}
    for raw_position, raw_codes in positions.items():
        position_text = str(raw_position).strip()
        if not position_text.isdigit() or int(position_text) <= 0:
            raise ValueError(f"некорректная позиция {raw_position!r}")
        position_text = str(int(position_text))
        if position_text in normalized:
            raise ValueError(f"позиция {position_text} повторяется")
        if not isinstance(raw_codes, list) or not raw_codes:
            raise ValueError(f"для позиции {position_text} ожидается непустой список")

        codes = []
        for raw_code in raw_codes:
            code = str(raw_code).strip()
            if not re.fullmatch(r"\d{4}", code):
                raise ValueError(
                    f"некорректный код {raw_code!r} для позиции {position_text}"
                )
            if code in codes:
                raise ValueError(
                    f"код {code} повторяется в позиции {position_text}"
                )
            previous_position = code_positions.get(code)
            if previous_position is not None:
                raise ValueError(
                    f"код {code} указан для позиций "
                    f"{previous_position} и {position_text}"
                )
            code_positions[code] = position_text
            codes.append(code)
        normalized[position_text] = tuple(codes)

    expected = set(expected_codes)
    actual = set(code_positions)
    if actual != expected:
        missing = ", ".join(sorted(expected - actual)) or "нет"
        extra = ", ".join(sorted(actual - expected)) or "нет"
        raise ValueError(
            f"состав кодов не совпадает со справочником "
            f"(отсутствуют: {missing}; лишние: {extra})"
        )

    return normalized


def load_additional_code_crosswalk(
    path,
    primary_code_to_position,
    valid_positions,
):
    """Загружает дополнительные коды, отсутствующие в основном справочнике."""
    try:
        catalog = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ValueError(f"файл {path.name} не найден") from exc
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"ошибка JSON в строке {exc.lineno}, столбце {exc.colno}"
        ) from exc

    if not isinstance(catalog, dict):
        raise ValueError("ожидается объект верхнего уровня")
    if catalog.get("expected_record_count") != 693:
        raise ValueError("expected_record_count должен быть равен 693")
    expected_group_counts = {"1": 4, "2": 653, "3": 28, "4": 8}
    if catalog.get("counts_by_group") != expected_group_counts:
        raise ValueError("counts_by_group не соответствует ожидаемому составу")

    records = catalog.get("records")
    if not isinstance(records, list) or len(records) != 693:
        raise ValueError("раздел records должен содержать 693 записи")

    normalized = {}
    actual_group_counts = {group: 0 for group in expected_group_counts}
    for index, record in enumerate(records, start=1):
        if not isinstance(record, dict):
            raise ValueError(f"запись {index} должна быть объектом")

        code = str(record.get("code", "")).strip()
        name = record.get("name")
        position = str(record.get("position", "")).strip()
        group = str(record.get("group", "")).strip()
        if not re.fullmatch(r"\d{4}", code):
            raise ValueError(f"некорректный код {code!r} в записи {index}")
        if code in normalized:
            raise ValueError(f"дополнительный код {code} повторяется")
        if code in primary_code_to_position:
            raise ValueError(f"дополнительный код {code} уже есть в основном справочнике")
        if not isinstance(name, str) or not name.strip():
            raise ValueError(f"для кода {code} отсутствует наименование")
        if position not in valid_positions:
            raise ValueError(f"для кода {code} указана неизвестная позиция {position!r}")
        if group not in actual_group_counts:
            raise ValueError(f"для кода {code} указана неизвестная группа {group!r}")

        normalized[code] = {
            "name": name.strip(),
            "position": position,
            "group": int(group),
        }
        actual_group_counts[group] += 1

    if actual_group_counts != expected_group_counts:
        raise ValueError("фактический состав групп не совпадает с counts_by_group")
    return normalized


try:
    POSITION_TO_RATE_CODES_2909 = load_position_crosswalk(
        POSITION_CROSSWALK_PATH,
        SUBSTANCE_NAMES,
    )
except ValueError as exc:
    st.error(f"Не удалось загрузить таблицу соответствий № 2909-р: {exc}.")
    st.stop()

CODE_TO_POSITION_2909 = {
    code: position
    for position, codes in POSITION_TO_RATE_CODES_2909.items()
    for code in codes
}

try:
    ADDITIONAL_SUBSTANCES_2909 = load_additional_code_crosswalk(
        ADDITIONAL_CODES_2909_PATH,
        CODE_TO_POSITION_2909,
        POSITION_TO_RATE_CODES_2909,
    )
except ValueError as exc:
    st.error(f"Не удалось загрузить дополнительные коды № 2909-р: {exc}.")
    st.stop()

ADDITIONAL_CODE_TO_POSITION_2909 = {
    code: record["position"]
    for code, record in ADDITIONAL_SUBSTANCES_2909.items()
}


def load_payment_rates(path, expected_positions):
    """Загружает и проверяет справочник ставок платы по годам."""
    try:
        catalog = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ValueError(f"файл {path.name} не найден") from exc
    except json.JSONDecodeError as exc:
        raise ValueError(f"ошибка JSON в строке {exc.lineno}, столбце {exc.colno}") from exc

    years = catalog.get("years") if isinstance(catalog, dict) else None
    if not isinstance(years, dict) or not years:
        raise ValueError("раздел years отсутствует или пуст")

    normalized_years = {}
    for year, config in years.items():
        year_text = str(year).strip()
        if not year_text:
            raise ValueError("обнаружен пустой год")
        if not isinstance(config, dict):
            raise ValueError(f"для {year_text} года ожидается объект")

        regulation = config.get("regulation")
        raw_rates = config.get("rates_by_position")
        if not isinstance(regulation, str) or not regulation.strip():
            raise ValueError(f"для {year_text} года не указан regulation")
        if not isinstance(raw_rates, dict) or not raw_rates:
            raise ValueError(
                f"для {year_text} года раздел rates_by_position отсутствует или пуст"
            )

        normalized_rates = {}
        for position, rate in raw_rates.items():
            position_text = str(position).strip()
            if not position_text.isdigit() or int(position_text) <= 0:
                raise ValueError(
                    f"некорректная позиция {position!r} для {year_text} года"
                )
            normalized_position = str(int(position_text))
            if normalized_position in normalized_rates:
                raise ValueError(
                    f"позиция {normalized_position} повторяется для {year_text} года"
                )
            if isinstance(rate, bool) or not isinstance(rate, (int, float)) or not np.isfinite(rate) or rate < 0:
                raise ValueError(
                    f"некорректная ставка для позиции {normalized_position} "
                    f"за {year_text} год"
                )
            normalized_rates[normalized_position] = float(rate)

        expected = set(expected_positions)
        actual = set(normalized_rates)
        if actual != expected:
            missing = ", ".join(sorted(expected - actual, key=int)) or "нет"
            extra = ", ".join(sorted(actual - expected, key=int)) or "нет"
            raise ValueError(
                f"состав позиций за {year_text} год не совпадает с № 2909-р "
                f"(отсутствуют: {missing}; лишние: {extra})"
            )

        normalized_years[year_text] = {
            "regulation": regulation.strip(),
            "rates_by_position": normalized_rates,
        }

    return normalized_years


try:
    PAYMENT_RATES_BY_YEAR = load_payment_rates(
        PAYMENT_RATES_PATH,
        POSITION_TO_RATE_CODES_2909,
    )
except ValueError as exc:
    st.error(f"Не удалось загрузить ставки платы: {exc}.")
    st.stop()

AVAILABLE_RATE_YEARS = sorted(
    PAYMENT_RATES_BY_YEAR,
    key=lambda year: int(year) if year.isdigit() else -1,
    reverse=True,
)


def find_rate_by_position(position, rates_by_position):
    """Ищет ставку по номеру пункта № 2409-р/2909-р."""
    if not position or pd.isna(position):
        return None

    position_text = str(position).strip()
    if not position_text.isdigit():
        return None
    return rates_by_position.get(str(int(position_text)))


def extract_position_from_substance(substance):
    """Извлекает позицию № 2909-р из начала наименования."""
    match = re.match(r'^\s*\((\d+)\)', substance)
    return str(int(match.group(1))) if match else None


def extract_code_from_substance(substance):
    """Извлекает четырехзначный код только из начала строки."""
    code_match = re.match(r'^\s*(\d{4})(?=\s|$)', substance)
    return code_match.group(1) if code_match else None


def resolve_rate_position(substance):
    """Определяет позицию напрямую или через код старого формата."""
    position = extract_position_from_substance(substance)
    if position is not None:
        return position if position in POSITION_TO_RATE_CODES_2909 else None

    code = extract_code_from_substance(substance)
    if code in CODE_TO_POSITION_2909:
        return CODE_TO_POSITION_2909[code]
    return ADDITIONAL_CODE_TO_POSITION_2909.get(code)

def calculate_payment(emission, rate, kvr, kpr):
    """Рассчитывает плату за выбросы с учетом коэффициентов Квр и Кпр"""
    if emission is None or pd.isna(emission):
        return None
    if rate is None or pd.isna(rate):
        return None

    try:
        return float(emission) * float(rate) * kvr * kpr
    except (TypeError, ValueError):
        return None


def calculate_selected_year_dataframe(
    dataframe,
    selected_year,
    rates_by_position,
    kvr,
    kpr,
):
    """Рассчитывает экран и первый лист по нормативу выбранного года."""
    result = dataframe.copy()
    norm_year = int(selected_year)
    result['Валовый выброс, т/год'] = pd.to_numeric(
        pd.Series(
            [
                annual_norms.get(norm_year)
                if isinstance(annual_norms, dict)
                else None
                for annual_norms in result['Нормативы по годам']
            ],
            index=result.index,
        ),
        errors='coerce',
    )
    result['Ставка платы, руб.'] = pd.to_numeric(
        pd.Series(
            [
                find_rate_by_position(position, rates_by_position)
                for position in result['Позиция 2909-р']
            ],
            index=result.index,
        ),
        errors='coerce',
    )
    result['Сумма платы, руб/год'] = pd.to_numeric(
        pd.Series(
            [
                calculate_payment(emission, rate, kvr, kpr)
                for emission, rate in zip(
                    result['Валовый выброс, т/год'],
                    result['Ставка платы, руб.'],
                )
            ],
            index=result.index,
        ),
        errors='coerce',
    )
    return result

def parse_emissions_file(uploaded_file):
    """
    Парсит файл с данными о выбросах в формате ПДВ
    Извлекает номер строки (столбец 1), наименование вещества (столбец 3) 
    и валовые выбросы т/год (столбец 6)
    """
    try:
        # Читаем CSV файл с разделителем ;
        df = pd.read_csv(
            uploaded_file, 
            sep=';', 
            encoding='utf-8',
            header=None,
            dtype=str,
            on_bad_lines='skip'
        )
        
        st.info(f"Файл загружен. Всего строк: {len(df)}")
        
        substances = []
        emissions = []
        codes = []
        positions_2909 = []
        row_numbers = []
        
        for idx, row in df.iterrows():
            if len(row) < 6:  # Пропускаем строки с недостаточным количеством колонок
                continue
            
            # Проверяем первую колонку - должен быть номер строки (число)
            first_col = str(row[0]).strip() if pd.notna(row[0]) else ""
            
            # Пропускаем строки, которые не начинаются с числа (заголовки, итоги и т.д.)
            if not first_col.isdigit():
                continue
            
            # Получаем номер строки
            row_num = int(first_col)
            
            # Наименование вещества - колонка 2 (индекс 1) - в файле это столбец 3
            substance = str(row[1]).strip() if len(row) > 1 and pd.notna(row[1]) else ""
            
            # Валовые выбросы - колонка 5 (индекс 5) - в файле это столбец 6
            emission_str = str(row[5]).strip() if len(row) > 5 and pd.notna(row[5]) else ""
            
            # Проверяем, что это действительно строка с данными (есть и вещество, и выбросы)
            if substance and emission_str and len(substance) > 3:
                # Очищаем название от кавычек если есть
                substance = substance.strip('"').strip("'")
                
                position_2909 = resolve_rate_position(substance)
                code = (
                    None
                    if extract_position_from_substance(substance) is not None
                    else extract_code_from_substance(substance)
                )
                
                # Заменяем запятую на точку для числа и удаляем пробелы
                emission_str = emission_str.replace(',', '.').replace(' ', '')
                
                try:
                    emission = float(emission_str)
                    substances.append(substance)
                    emissions.append(emission)
                    codes.append(code)
                    positions_2909.append(position_2909)
                    row_numbers.append(row_num)
                except ValueError:
                    # Если не удалось преобразовать в число, пропускаем
                    continue
        
        if substances:
            # Создаем DataFrame с нужными колонками
            result_df = pd.DataFrame({
                'Позиция 2909-р': positions_2909,
                'Код вещества': codes,
                'Наименование вещества': substances,
                'Валовый выброс, т/год': emissions
            })
            return result_df
        else:
            st.warning("Не удалось найти данные в файле. Проверьте формат файла.")
            return None
            
    except Exception as e:
        st.error(f"Ошибка при чтении файла: {str(e)}")
        return None

@dataclass
class XlsParseResult:
    """Результат парсинга без передачи технических деталей в аналитику."""

    dataframe: pd.DataFrame | None
    error_category: str | None = None
    user_message: str | None = None
    years: tuple[int, ...] = ()
    warnings: tuple[str, ...] = ()

    @property
    def success(self):
        return self.dataframe is not None and not self.dataframe.empty


def _empty_emissions_dataframe():
    return pd.DataFrame(columns=[
        'Позиция 2909-р',
        'Код вещества',
        'Наименование вещества',
        'Валовый выброс, т/год',
        'Нормативы по годам',
    ])


OBJECT_FORMAT_TITLE = (
    'нормативы выбросов загрязняющих веществ от стационарных изав '
    'в атмосферный воздух по объекту онв'
)
SOURCE_FORMAT_TITLE = (
    'нормативы выбросов загрязняющих веществ в атмосферный воздух '
    'по конкретным стационарным источникам выбросов и загрязняющим веществам'
)


BINARY_XLS_SIGNATURE = bytes.fromhex('D0 CF 11 E0 A1 B1 1A E1')
ANNUAL_PERIOD_LENGTH = 8
MODIFIED_STRUCTURE_MESSAGE = (
    'Файл не соответствует ожидаемой структуре. Вероятно, столбцы были '
    'перемещены или удалены. Загрузите исходный файл из ПДВ-Эколог без изменений.'
)


class XlsStructureError(ValueError):
    """Структура выгрузки отличается от поддерживаемого шаблона."""


def _normalize_xls_text(value):
    """Нормализует текст SpreadsheetML для устойчивого поиска заголовков."""
    return ' '.join(str(value or '').split()).strip()


def _spreadsheetml_rows(worksheet, namespace):
    """Возвращает строки SpreadsheetML с учетом разреженных индексов ячеек."""
    rows = []
    ss_ns = 'urn:schemas-microsoft-com:office:spreadsheet'
    for row_elem in worksheet.findall('.//ss:Row', namespace):
        row_dict = {}
        col_idx = 1
        for cell in row_elem.findall('ss:Cell', namespace):
            explicit = cell.get(f'{{{ss_ns}}}Index')
            if explicit:
                col_idx = int(explicit)
            data_elem = cell.find('ss:Data', namespace)
            row_dict[col_idx] = data_elem.text if data_elem is not None else None
            col_idx += 1
        rows.append(row_dict)
    return rows


def _binary_xls_rows(raw_content):
    """Читает первый лист бинарного Excel 97–2003 из памяти."""
    workbook = xlrd.open_workbook(file_contents=raw_content)
    if workbook.nsheets == 0:
        return []

    worksheet = workbook.sheet_by_index(0)
    rows = []
    for row_idx in range(worksheet.nrows):
        row_dict = {}
        for col_idx in range(worksheet.ncols):
            cell = worksheet.cell(row_idx, col_idx)
            if cell.ctype != xlrd.XL_CELL_EMPTY:
                row_dict[col_idx + 1] = cell.value
        rows.append(row_dict)
    return rows

def _detect_emissions_xls_format(rows):
    """Определяет вариант выгрузки ПДВ-Эколог по внутреннему заголовку."""
    workbook_text = ' '.join(
        _normalize_xls_text(value).lower()
        for row in rows
        for value in row.values()
        if value
    )
    if SOURCE_FORMAT_TITLE in workbook_text:
        return 'sources'
    if OBJECT_FORMAT_TITLE in workbook_text:
        return 'object'
    return None


def _parse_xls_number(value):
    normalized = _normalize_xls_text(value).replace(' ', '').replace(',', '.')
    return float(normalized)


def _detect_annual_layout(rows, xls_format):
    """Проверяет позиции и состав всех восьмилетних блоков."""
    first_header_column = 5 if xls_format == 'object' else 4
    expected_header_columns = tuple(
        first_header_column + offset * 3
        for offset in range(ANNUAL_PERIOD_LENGTH)
    )
    detected_years = None
    header_rows_found = 0

    for row_index, row in enumerate(rows):
        year_cells = []
        for column_index, value in row.items():
            match = re.search(
                r'(?<!\d)(20\d{2})\s*год',
                _normalize_xls_text(value).lower(),
            )
            if match:
                year_cells.append((int(match.group(1)), column_index))
        if len(year_cells) < 2:
            continue

        header_rows_found += 1
        years_by_column = sorted(
            (column_index, year)
            for year, column_index in year_cells
        )
        header_columns = tuple(column for column, _year in years_by_column)
        years = tuple(year for _column, year in years_by_column)
        expected_years = tuple(range(years[0], years[0] + ANNUAL_PERIOD_LENGTH))
        if (
            len(year_cells) != ANNUAL_PERIOD_LENGTH
            or header_columns != expected_header_columns
            or years != expected_years
            or (detected_years is not None and years != detected_years)
        ):
            raise XlsStructureError

        if row_index + 1 >= len(rows):
            raise XlsStructureError
        units_row = rows[row_index + 1]
        for header_column in expected_header_columns:
            units = tuple(
                _normalize_xls_text(units_row.get(header_column + offset))
                .lower()
                .replace(' ', '')
                for offset in range(3)
            )
            if (
                units[0] != 'г/с'
                or units[1] not in {'т/г', 'т/год'}
                or units[2] != 'пдв/врв'
            ):
                raise XlsStructureError

        detected_years = years

    if header_rows_found == 0 or detected_years is None:
        raise XlsStructureError

    tons_columns = {
        year: header_column + 1
        for year, header_column in zip(detected_years, expected_header_columns)
    }
    return detected_years, tons_columns

def _parse_annual_norms(row, substance, years, tons_columns, warnings):
    annual_norms = {}
    for year in years:
        value = row.get(tons_columns[year])
        try:
            annual_norms[year] = _parse_xls_number(value)
        except (TypeError, ValueError):
            annual_norms[year] = None
            warnings.append(
                f'{substance}: отсутствует или некорректен норматив за {year} год'
            )
    return annual_norms


def _build_emissions_dataframe(records, years):
    if not records:
        return _empty_emissions_dataframe()

    substances, emissions, codes = [], [], []
    positions_2909, annual_values = [], []
    first_year = years[0]
    for substance, annual_norms in records:
        position_2909 = resolve_rate_position(substance)
        code = (
            None
            if extract_position_from_substance(substance) is not None
            else extract_code_from_substance(substance)
        )
        substances.append(substance)
        emissions.append(annual_norms.get(first_year))
        codes.append(code)
        positions_2909.append(position_2909)
        annual_values.append(annual_norms)

    return pd.DataFrame({
        'Позиция 2909-р': positions_2909,
        'Код вещества': codes,
        'Наименование вещества': substances,
        'Валовый выброс, т/год': emissions,
        'Нормативы по годам': annual_values,
    })


def _parse_object_format_rows(rows, years, tons_columns, warnings):
    """Разбирает сводную таблицу нормативов по объекту ОНВ."""
    records = []
    for row in rows:
        row_number = row.get(1)
        is_number = (
            isinstance(row_number, (int, float))
            and float(row_number).is_integer()
        )
        if not is_number and not _normalize_xls_text(row_number).isdigit():
            continue
        substance = _normalize_xls_text(row.get(2))
        if len(substance) < 4:
            continue
        annual_norms = _parse_annual_norms(
            row, substance, years, tons_columns, warnings
        )
        records.append((substance, annual_norms))
    return records


def _parse_source_format_rows(rows, years, tons_columns, warnings):
    """Разбирает таблицу по источникам, используя только строки «Всего по ЗВ»."""
    records = []
    current_substance = None
    substance_heading = 'наименование и код загрязняющего вещества'
    for row in rows:
        first_cell = _normalize_xls_text(row.get(1)).lower()
        if substance_heading in first_cell:
            substance = _normalize_xls_text(row.get(5))
            current_substance = substance if len(substance) >= 4 else None
            continue
        total_label = _normalize_xls_text(row.get(2)).lower()
        if current_substance and total_label == 'всего по зв':
            annual_norms = _parse_annual_norms(
                row, current_substance, years, tons_columns, warnings
            )
            records.append((current_substance, annual_norms))
            current_substance = None
    return records


def parse_emissions_xls(uploaded_file):
    """Парсит XML и бинарные XLS-файлы из ПДВ-Эколог."""
    try:
        raw_content = uploaded_file.read()
    except OSError:
        return XlsParseResult(
            None,
            'read_error',
            "Не удалось прочитать XLS-файл. Повторно выгрузите его из ПДВ-Эколог.",
        )

    stripped_content = raw_content.lstrip()
    if stripped_content.startswith(b'<?xml') or stripped_content.startswith(b'<Workbook'):
        try:
            content = raw_content.decode('utf-8-sig')
            tree = ET.fromstring(content)
        except (UnicodeError, ET.ParseError):
            return XlsParseResult(
                None,
                'invalid_xml',
                "Не удалось прочитать XML-структуру XLS-файла. "
                "Повторно выгрузите файл из ПДВ-Эколог без изменений.",
            )

        ss_ns = 'urn:schemas-microsoft-com:office:spreadsheet'
        ns = {'ss': ss_ns}
        worksheets = tree.findall('.//ss:Worksheet', ns)
        if not worksheets:
            return XlsParseResult(
                None,
                'no_worksheets',
                "В XLS-файле не найдено листов.",
            )
        rows = _spreadsheetml_rows(worksheets[0], ns)
    elif raw_content.startswith(BINARY_XLS_SIGNATURE):
        try:
            rows = _binary_xls_rows(raw_content)
        except Exception:
            return XlsParseResult(
                None,
                'invalid_binary_xls',
                "Не удалось прочитать бинарный XLS-файл. "
                "Повторно выгрузите файл из ПДВ-Эколог без изменений.",
            )
        if not rows:
            return XlsParseResult(
                None,
                'no_worksheets',
                "В XLS-файле не найдено листов.",
            )
    else:
        return XlsParseResult(
            None,
            'unsupported_format',
            "Файл не является поддерживаемым XLS. Загрузите исходную выгрузку "
            "из ПДВ-Эколог без изменения формата.",
        )

    try:
        xls_format = _detect_emissions_xls_format(rows)
        warnings = []
        if xls_format in {'object', 'sources'}:
            try:
                years, tons_columns = _detect_annual_layout(rows, xls_format)
            except XlsStructureError:
                return XlsParseResult(
                    None,
                    'modified_structure',
                    MODIFIED_STRUCTURE_MESSAGE,
                )

        if xls_format == 'object':
            records = _parse_object_format_rows(
                rows, years, tons_columns, warnings
            )
        elif xls_format == 'sources':
            records = _parse_source_format_rows(
                rows, years, tons_columns, warnings
            )
        else:
            return XlsParseResult(
                None,
                'unsupported_format',
                'Не удалось определить вид таблицы. Загрузите исходную выгрузку '
                'нормативов по объекту ОНВ или по конкретным стационарным источникам.',
            )

        if not records:
            return XlsParseResult(_empty_emissions_dataframe(), 'no_data')
        return XlsParseResult(
            _build_emissions_dataframe(records, years),
            years=years,
            warnings=tuple(warnings),
        )
    except Exception:
        return XlsParseResult(
            None,
            'unexpected_error',
            "Не удалось обработать данные XLS-файла. "
            "Повторно выгрузите файл из ПДВ-Эколог без изменений.",
        )

def format_dataframe_for_display(df):
    """Форматирует DataFrame для отображения с округлением"""
    display_df = df.copy()
    
    # None для ненайденных ставок преобразуется в NaN перед округлением.
    display_df['Валовый выброс, т/год'] = pd.to_numeric(
        display_df['Валовый выброс, т/год'], errors='coerce'
    ).round(6)
    display_df['Ставка платы, руб.'] = pd.to_numeric(
        display_df['Ставка платы, руб.'], errors='coerce'
    ).round(2)
    display_df['Сумма платы, руб/год'] = pd.to_numeric(
        display_df['Сумма платы, руб/год'], errors='coerce'
    ).round(2)

    display_df.index = range(1, len(display_df) + 1)
    return display_df

def add_total_row(df):
    """Добавляет строку с итогами в DataFrame"""
    total_row = pd.DataFrame({
        'Наименование вещества': ['ИТОГО:'],
        'Валовый выброс, т/год': [df['Валовый выброс, т/год'].sum()],
        'Ставка платы, руб.': [np.nan],
        'Сумма платы, руб/год': [df['Сумма платы, руб/год'].sum()]
    })
    return pd.concat([df, total_row], ignore_index=True)


def build_single_year_export_dataframe(df):
    """Формирует первый лист с расчетом за выбранный год."""
    columns = [
        'Наименование вещества',
        'Валовый выброс, т/год',
        'Ставка платы, руб.',
        'Сумма платы, руб/год',
    ]
    export_df = add_total_row(df[columns].copy())
    return format_dataframe_for_display(export_df)


def resolve_rate_year(norm_year, rates_by_year):
    """Возвращает год ставки, используя последний доступный для будущих лет."""
    year_text = str(norm_year)
    if year_text in rates_by_year:
        return year_text

    numeric_years = sorted(
        int(year)
        for year in rates_by_year
        if str(year).isdigit()
    )
    if numeric_years and int(norm_year) > numeric_years[-1]:
        return str(numeric_years[-1])
    return None


def build_yearly_export_dataframe(df, years, rates_by_year, kvr, kpr):
    """Формирует матрицу «норматив — ставка — сумма» для восьми лет."""
    export_data = {
        'Наименование вещества': df['Наименование вещества'].tolist(),
    }

    for year in years:
        rate_year = resolve_rate_year(year, rates_by_year)
        rates = (
            rates_by_year[rate_year]['rates_by_position']
            if rate_year is not None
            else {}
        )
        norms = pd.Series(
            [
                annual_norms.get(year)
                if isinstance(annual_norms, dict)
                else None
                for annual_norms in df['Нормативы по годам']
            ],
            index=df.index,
            dtype='float64',
        )
        year_rates = pd.Series(
            [
                find_rate_by_position(position, rates)
                for position in df['Позиция 2909-р']
            ],
            index=df.index,
            dtype='float64',
        )
        payments = pd.Series(
            [
                calculate_payment(norm, rate, kvr, kpr)
                if pd.notna(norm)
                else None
                for norm, rate in zip(norms, year_rates)
            ],
            index=df.index,
            dtype='float64',
        )
        export_data[f'Норматив {year}, т/год'] = norms.tolist()
        export_data[f'Ставка {year}, руб./т'] = year_rates.tolist()
        export_data[f'Сумма платы {year}, руб/год'] = payments.tolist()

    yearly_df = pd.DataFrame(export_data)
    total_row = {
        'Наименование вещества': 'ИТОГО:',
    }
    for year in years:
        norm_column = f'Норматив {year}, т/год'
        rate_column = f'Ставка {year}, руб./т'
        payment_column = f'Сумма платы {year}, руб/год'
        norm_values = pd.to_numeric(yearly_df[norm_column], errors='coerce')
        payment_values = pd.to_numeric(yearly_df[payment_column], errors='coerce')
        total_row[norm_column] = norm_values.sum(skipna=True)
        total_row[rate_column] = np.nan
        total_row[payment_column] = payment_values.sum(skipna=True)

    return pd.concat(
        [yearly_df, pd.DataFrame([total_row])],
        ignore_index=True,
    )


def create_excel_output(
    single_year_df,
    yearly_df,
    years,
    rates_by_year,
    selected_year,
):
    """Создаёт книгу с однолетним и полным восьмилетним расчётом."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        single_year_df.to_excel(
            writer,
            sheet_name='Расчет платы',
            index=False,
        )
        yearly_df.to_excel(
            writer,
            sheet_name='Расчёт по годам',
            index=False,
        )

        single_sheet = writer.sheets['Расчет платы']
        yearly_sheet = writer.sheets['Расчёт по годам']
        payment_fill = PatternFill('solid', fgColor='E4F1E8')
        thin_side = Side(style='thin', color='D9D9D9')
        grid_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )
        total_border = Border(
            left=thin_side,
            right=thin_side,
            top=Side(style='medium', color='315B7D'),
            bottom=thin_side,
        )

        sheet_configs = (
            (single_sheet, single_year_df, {4}, {3}),
            (
                yearly_sheet,
                yearly_df,
                {4 + offset * 3 for offset, _year in enumerate(years)},
                {3 + offset * 3 for offset, _year in enumerate(years)},
            ),
        )
        for worksheet, dataframe, payment_columns, rate_columns in sheet_configs:
            last_row = len(dataframe) + 1
            last_data_row = max(1, last_row - 1)
            last_column = len(dataframe.columns)
            worksheet.freeze_panes = 'B2'
            worksheet.auto_filter.ref = (
                f'A1:{get_column_letter(last_column)}{last_data_row}'
            )
            worksheet.sheet_view.showGridLines = False
            worksheet.row_dimensions[1].height = 42

            for row_index in range(1, last_row + 1):
                for column_index in range(1, last_column + 1):
                    cell = worksheet.cell(row=row_index, column=column_index)
                    cell.border = grid_border
                    if column_index in payment_columns:
                        cell.fill = payment_fill

            for column_index in range(1, last_column + 1):
                cell = worksheet.cell(row=1, column=column_index)
                cell.font = Font(bold=True, color='1F2937')
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True,
                )

            worksheet.column_dimensions['A'].width = 50
            for column_index in range(2, last_column + 1):
                worksheet.column_dimensions[
                    get_column_letter(column_index)
                ].width = 20

            for row_index in range(2, last_row + 1):
                for column_index in range(2, last_column + 1):
                    worksheet.cell(
                        row=row_index,
                        column=column_index,
                    ).number_format = (
                        '0.000000'
                        if column_index not in rate_columns | payment_columns
                        else '0.00'
                    )

            for column_index in range(1, last_column + 1):
                cell = worksheet.cell(row=last_row, column=column_index)
                cell.font = Font(bold=True)
                cell.border = total_border

        single_note_row = len(single_year_df) + 3
        single_sheet.merge_cells(
            start_row=single_note_row,
            start_column=1,
            end_row=single_note_row,
            end_column=len(single_year_df.columns),
        )
        single_note_cell = single_sheet.cell(row=single_note_row, column=1)
        single_note_cell.value = (
            f'Примечание: расчёт выполнен по нормативам и ставкам '
            f'за {selected_year} год.'
        )
        single_note_cell.font = Font(italic=True, color='595959')
        single_note_cell.alignment = Alignment(wrap_text=True, vertical='top')
        single_sheet.row_dimensions[single_note_row].height = 28

        fallback_years = [
            year
            for year in years
            if resolve_rate_year(year, rates_by_year) != str(year)
            and resolve_rate_year(year, rates_by_year) is not None
        ]
        if fallback_years:
            fallback_rate_year = resolve_rate_year(fallback_years[0], rates_by_year)
            years_text = (
                str(fallback_years[0])
                if len(fallback_years) == 1
                else f'{fallback_years[0]}–{fallback_years[-1]}'
            )
            note_row = len(yearly_df) + 3
            last_column = len(yearly_df.columns)
            yearly_sheet.merge_cells(
                start_row=note_row,
                start_column=1,
                end_row=note_row,
                end_column=last_column,
            )
            note_cell = yearly_sheet.cell(row=note_row, column=1)
            note_cell.value = (
                f'Примечание: для расчёта платы за {years_text} годы условно '
                f'применены ставки платы, установленные на {fallback_rate_year} '
                'год, в связи с отсутствием утверждённых ставок на указанный '
                'период.'
            )
            note_cell.font = Font(italic=True, color='595959')
            note_cell.alignment = Alignment(wrap_text=True, vertical='top')
            yearly_sheet.row_dimensions[note_row].height = 32

    output.seek(0)
    return output

def handle_emissions_file_change():
    """Фиксирует новый выбор файла и сбрасывает кэш предыдущего разбора."""
    st.session_state['_upload_revision'] = st.session_state.get('_upload_revision', 0) + 1
    st.session_state.pop('_parsed_upload_key', None)
    st.session_state.pop('_parsed_upload_result', None)
    if st.session_state.get('emissions_file') is not None:
        queue_event('file_uploaded')


def handle_excel_download():
    queue_event('excel_download')

# Загрузка файла с выбросами
st.markdown(
    """
    <div class="eco-important-notice">
      <span class="eco-important-icon" aria-hidden="true">!</span>
      <div><strong>ВАЖНО</strong><br>
      <strong>Загрузите файл с</strong> таблицей <strong>нормативов выбросов по объекту ОНВ в целом</strong> или
      <strong>Нормативы по стационарным источникам выбросов</strong> из ПДВ-Эколог
      (сохранив его <strong>без изменений</strong> как Файл MS Excel (*.xls)).
      Не двигайте столбцы, не удаляйте строки в нем.
      <div class="eco-supported-formats"><strong>Поддерживаемые форматы:</strong><br>
      <span class="eco-format-icon eco-format-ok" aria-hidden="true">✓</span> Документ Excel (XML)<br>
      <span class="eco-format-icon eco-format-ok" aria-hidden="true">✓</span> Документ Excel 97/2000/XP<br>
      <span class="eco-format-icon eco-format-no" aria-hidden="true">×</span> Excel (OLE) / PDF / Word / HTML — пока не поддерживаются</div>
      <div class="eco-privacy-note">Загруженный файл используется только для расчета и не сохраняется.<br>
      Для анализа использования применяется Google Analytics. Имена и содержимое файлов,
      данные расчёта и суммы в аналитику не передаются.</div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)
emissions_file = st.file_uploader(
    "Загрузите исходный файл ПДВ в формате .xls",
    type=['xls'],
    key='emissions_file',
    on_change=handle_emissions_file_change,
    help="Загрузите исходный файл ПДВ в формате .xls без изменений.",
)

# Настройка коэффициентов в сайдбаре
with st.sidebar:
    st.markdown(
        """
        <div class="eco-sidebar-brand">
          <div class="eco-logo" aria-hidden="true"><span></span></div>
          <div><strong>Ecolytica</strong></div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    render_sidebar_about()
    st.markdown(
        '<div class="eco-section-label eco-sidebar-label">Параметры</div>',
        unsafe_allow_html=True,
    )
    st.header("Настройки расчёта")

    current_rate_year = str(datetime.now().year)
    default_rate_year_index = (
        AVAILABLE_RATE_YEARS.index(current_rate_year)
        if current_rate_year in AVAILABLE_RATE_YEARS
        else 0
    )
    selected_year = st.selectbox(
        "Год ставок платы",
        options=AVAILABLE_RATE_YEARS,
        index=default_rate_year_index,
        help="Список формируется автоматически из data/payment_rates.json",
    )
    selected_rate_config = PAYMENT_RATES_BY_YEAR[selected_year]
    selected_rates_by_position = selected_rate_config["rates_by_position"]
    st.caption(selected_rate_config["regulation"])

    st.subheader("Коэффициенты для расчета платы")
    
    # Коэффициент Квр (выпадающий список 1 или 25)
    kvr = st.selectbox(
        "Коэффициент к ставке платы за выброс в пределах ВРВ (Квр)",
        options=[1, 25],
        index=0,
        help="Коэффициент для выбросов в пределах временно разрешенных выбросов (ВРВ)"
    )
    
    # Коэффициент Кпр (выпадающий список 1, 25 или 100)
    kpr = st.selectbox(
        "Коэффициент к ставке платы за выброс сверх ВРВ, НДВ, ТН (Кпр)",
        options=[1, 25, 100],
        index=0,
        help="Коэффициент для выбросов сверх установленных нормативов"
    )
    
        
    # Поиск по коду
    with st.expander("Поиск ставки по коду"):
        search_code = st.text_input("Введите код вещества (4 цифры)")
        if search_code:
            code_padded = search_code.zfill(4)
            position = CODE_TO_POSITION_2909.get(code_padded)
            additional_record = ADDITIONAL_SUBSTANCES_2909.get(code_padded)
            if position is None and additional_record is not None:
                position = additional_record["position"]
            rate = find_rate_by_position(position, selected_rates_by_position)
            if rate is not None:
                name = SUBSTANCE_NAMES.get(
                    code_padded,
                    additional_record["name"]
                    if additional_record is not None
                    else "Наименование не указано",
                )
                st.success(
                    f"**{code_padded}**, пункт № {position}: {name[:100]}..."
                )
                st.metric("Ставка", f"{rate:,.2f} руб/т")
            else:
                st.warning("Код не найден")

# Основная логика расчета
if emissions_file is not None:
    upload_key = st.session_state.get('_upload_revision', 0)
    with st.spinner('Обработка файла с выбросами...'):
        if st.session_state.get('_parsed_upload_key') != upload_key:
            emissions_file.seek(0)
            parse_result = parse_emissions_xls(emissions_file)
            st.session_state['_parsed_upload_key'] = upload_key
            st.session_state['_parsed_upload_result'] = parse_result
            if parse_result.success:
                queue_event('xls_parse_success')
            else:
                queue_event(
                    'xls_parse_failed',
                    error_category=parse_result.error_category or 'unexpected_error',
                )
        else:
            parse_result = st.session_state['_parsed_upload_result']

        df_result = (
            parse_result.dataframe.copy()
            if parse_result.dataframe is not None
            else None
        )
        if parse_result.user_message:
            if parse_result.error_category == 'modified_structure':
                st.warning(parse_result.user_message)
            else:
                st.error(parse_result.user_message)

        if parse_result.success:
            df_result = calculate_selected_year_dataframe(
                df_result,
                selected_year,
                selected_rates_by_position,
                kvr,
                kpr,
            )
            unresolved_substances = df_result.loc[
                df_result['Ставка платы, руб.'].isna(),
                'Наименование вещества',
            ].tolist()
            
            # Считаем итоги
            total_payment = df_result['Сумма платы, руб/год'].sum(skipna=True)
            total_emission = df_result['Валовый выброс, т/год'].sum()
            substances_with_rate = df_result['Ставка платы, руб.'].notna().sum()
            
            st.success(f"Обработка завершена: найдено записей — {len(df_result)}")
            if parse_result.warnings:
                st.warning(
                    "В годовых нормативах найдено пропусков или некорректных "
                    f"значений: {len(parse_result.warnings)}. "
                    "Они оставлены пустыми и не включаются в соответствующие "
                    "годовые суммы."
                )
            if unresolved_substances:
                substances_text = "; ".join(unresolved_substances[:10])
                suffix = "…" if len(unresolved_substances) > 10 else ""
                st.warning(
                    "Не найдена позиция или ставка № 2409-р/2909-р для: "
                    f"{substances_text}{suffix}"
                )
            
            # Отображаем итоговую сумму
            st.markdown('<div class="eco-section-label eco-results-label">Результаты расчёта</div>', unsafe_allow_html=True)
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric(
                    "Итого плата",
                    f"{total_payment:.2f} руб" if pd.notna(total_payment) else "0 руб"
                )
            with col2:
                st.metric("Всего веществ", len(df_result))
            with col3:
                st.metric("Найдено ставок", f"{substances_with_rate} из {len(df_result)}")
            
            # Отображаем выбранные коэффициенты
            st.info(
                f"**Год норматива и ставок:** {selected_year}. "
                f"**Выбранные коэффициенты:** Квр = {kvr}, Кпр = {kpr}"
            )
            st.caption(
                f"Расчет в таблице выполнен по нормативу {selected_year} года. "
                "Расчет по годам доступен в результатах Excel."
            )
            st.markdown('<div class="eco-table-label">Расчёт по веществам</div>', unsafe_allow_html=True)
            
            # Таблица результатов
            display_columns = ['Наименование вещества', 'Валовый выброс, т/год',
                              'Ставка платы, руб.', 'Сумма платы, руб/год']
            display_df = df_result[display_columns].copy()
            display_df = format_dataframe_for_display(display_df)
            st.dataframe(
                display_df,
                width="stretch",
                height=600,
                column_config={
                    'Валовый выброс, т/год': st.column_config.NumberColumn(
                        format='%.6f'
                    ),
                    'Ставка платы, руб.': st.column_config.NumberColumn(
                        format='%.2f'
                    ),
                    'Сумма платы, руб/год': st.column_config.NumberColumn(
                        format='%.2f'
                    ),
                },
            )

            st.markdown(
                """
                <div class="eco-social-cta">
                  Если инструмент сэкономил вам время — подпишитесь на мой канал <strong>ВК или Telegram</strong>.
                  <p>Там показываю, как упростить жизнь экологу-проектировщику.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
            telegram_col, vk_col = st.columns(2)
            with telegram_col:
                render_social_button(
                    "Подписаться в Telegram",
                    "https://t.me/ecology_start",
                    "telegram",
                )
            with vk_col:
                render_social_button(
                    "Вступить в ВК",
                    "https://vk.ru/ecolytica",
                    "vk",
                )
            
            # НОВЫЙ БЛОК: Кнопка для скачивания в формате Excel с итоговой строкой
            st.subheader("Экспорт результатов")

            single_year_export_df = build_single_year_export_dataframe(
                df_result
            )
            yearly_export_df = build_yearly_export_dataframe(
                df_result,
                parse_result.years,
                PAYMENT_RATES_BY_YEAR,
                kvr,
                kpr,
            )
            output = create_excel_output(
                single_year_export_df,
                yearly_export_df,
                parse_result.years,
                PAYMENT_RATES_BY_YEAR,
                selected_year,
            )

            # Подготавливаем файл для скачивания
            output.seek(0)
            
            # Генерируем имя файла с текущей датой
            current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"raschet_platy_{current_date}.xlsx"
            
            st.download_button(
                label="Скачать результаты в Excel",
                data=output,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                on_click=handle_excel_download,
                width="stretch",
            )
                
        elif parse_result.error_category == 'no_data':
            st.error(
                "Данных не найдено, либо загруженный файл сохранен с изменениями. "
                "Повторно выгрузите файл из программы ПДВ и сохраните его без изменений."
            )

# Инструкция в сайдбаре
with st.sidebar:
    st.divider()
    st.header("Инструкция")
    
    st.markdown("""
    ### Как пользоваться:
    1. **Загрузите файл с нормативами выбросов** из программы ПДВ (сохранив **без изменений** как Файл MS Excel (*.xls))
    2. Выберите **год ставок**, коэффициенты **Квр** и **Кпр** из выпадающих списков
    3. Приложение автоматически найдёт ставки по коду вещества и рассчитает плату
    4. Скачайте результаты в формате Excel
    
    ### Формула расчета:
    **Плата = Выбросы (т/год) × Ставка × Квр × Кпр**
    
    где:
    - **Квр** - коэффициент для выбросов в пределах ВРВ (1 или 25)
    - **Кпр** - коэффициент для выбросов сверх нормативов (1, 25 или 100)
    
       """)

st.markdown('<div class="eco-feedback-label">Обратная связь</div>', unsafe_allow_html=True)
render_feedback()
flush_events()

st.markdown(
    '<footer class="eco-footer">Проект <strong>«Экология без ручной рутины | Ecolytica»</strong></footer>',
    unsafe_allow_html=True,
)
